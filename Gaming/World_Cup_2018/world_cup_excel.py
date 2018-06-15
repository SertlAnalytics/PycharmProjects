"""
Description: This module contains the excel class for FIFA World cup predictor.
Author: Josef Sertl
Copyright: SERTL Analytics, https://sertl-analytics.com
Date: 2018-06-15
"""

from world_cup_constants import FC
import openpyxl


class WorldCupExcel:
    def __init__(self, file_path: str, work_sheet_name: str):
        self.__file_path = file_path
        self.__work_book = openpyxl.load_workbook(self.__file_path)
        self.__work_sheet = self.__work_book[work_sheet_name]
        self.__col_dic = self.__get_column_dic__()
        self.__match_number_row_dic = self.__get_match_number_dic__()

    def update_match(self, match_number, goal_1: int, goal_2: int, goal_1_penalty = 0, goal_2_penalty = 0):
        row_match = self.__match_number_row_dic[match_number]
        if self.get_value(row_match, self.__col_dic[FC.STATUS]) == 'open':
            self.set_value(row_match, self.__col_dic[FC.GOALS_HOME], goal_1)
            self.set_value(row_match, self.__col_dic[FC.GOALS_AWAY], goal_2)
            self.set_value(row_match, self.__col_dic[FC.HOME_PENALTIES], goal_1_penalty)
            self.set_value(row_match, self.__col_dic[FC.AWAY_PENALTIES], goal_2_penalty)
            self.set_value(row_match, self.__col_dic[FC.STATUS], 'completed')
            self.__work_book.save(self.__file_path)

    def get_value(self, row: int, column: int):
        cell = self.__work_sheet.cell(row, column)
        return cell.value

    def set_value(self, row: int, column: int, value):
        cell = self.__work_sheet.cell(row, column)
        cell.value = value

    def __get_column_dic__(self):
        col_dic = {}
        for col in range(1, 100):
            cell = self.__work_sheet.cell(1, col)
            if cell.value is None:
                break
            col_dic[cell.value] = col
        return col_dic

    def __get_match_number_dic__(self):
        match_number_dic = {}
        for row in range(1, 100):
            cell = self.__work_sheet.cell(row, self.__col_dic[FC.MATCH_NUMBER])
            if cell.value is None:
                break
            match_number_dic[cell.value] = row
        return match_number_dic

